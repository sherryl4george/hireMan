import openpyxl as opx
from mako.filters import trim

from app.StudentNewHire import StudentNewHire

def process_bulkupload(filename, fileurl=None):
    wb = opx.load_workbook(filename=filename, data_only=True)
    ws = wb['GA_Template']
    for row in range(2,ws.max_row+1):
        values = []
        for col in range(1,ws.max_column+1):
            value = ws.cell(row,col).value
            values.append(value)

        item = StudentNewHire(*values)
        if item.validate():
            print("Success")
        else:
            print("Failure")






